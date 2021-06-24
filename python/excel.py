from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
print(wb.sheetnames)

for row in wb.active.values:
    for value in row:
        print(value)